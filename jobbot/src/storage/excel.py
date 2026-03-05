"""Excel tracker using openpyxl — upsert rows and maintain status transitions."""

import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH = Path(__file__).resolve().parents[2] / "data" / "applications.xlsx"

COLUMNS = [
    "app_id", "source", "company", "role_title", "role_family", "job_url",
    "location", "match_score", "resume_version", "status", "policy",
    "completeness_pct", "missing_fields_count", "date_discovered",
    "date_submitted", "submission_proof", "proof_path", "next_action_due", "notes"
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

STATUS_COLORS = {
    "DISCOVERED": "E2EFDA",
    "MAPPABLE": "DDEBF7",
    "FILLING": "FFF2CC",
    "FILLED_AWAITING_SUBMIT": "FFEB9C",
    "NEEDS_USER_DATA": "FCE4D6",
    "READY_TO_APPLY": "BDD7EE",
    "APPLYING": "FFF2CC",
    "APPLIED": "C6EFCE",
    "SUBMITTED": "C6EFCE",
    "NEEDS_HUMAN": "FFC7CE",
    "REJECTED": "D9D9D9",
    "INTERVIEW": "B4C6E7",
}

VALID_TRANSITIONS = {
    "DISCOVERED": ["READY_TO_APPLY", "NEEDS_HUMAN"],
    "READY_TO_APPLY": ["APPLYING", "NEEDS_HUMAN"],
    "APPLYING": ["SUBMITTED", "NEEDS_HUMAN"],
    "SUBMITTED": ["INTERVIEW", "REJECTED", "NEEDS_HUMAN"],
    "NEEDS_HUMAN": ["READY_TO_APPLY", "APPLYING", "SUBMITTED"],
    "INTERVIEW": ["REJECTED"],
    "REJECTED": [],
}


def _ensure_workbook(path: Optional[Path] = None) -> tuple[Workbook, Path]:
    """Load existing workbook or create new one with headers."""
    p = path or EXCEL_PATH
    p.parent.mkdir(parents=True, exist_ok=True)

    if p.exists():
        wb = load_workbook(str(p))
        ws = wb.active
        # Verify headers match
        existing_headers = [cell.value for cell in ws[1]]
        if existing_headers != COLUMNS:
            # Re-create headers if schema changed
            for col_idx, col_name in enumerate(COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        # Set reasonable column widths
        widths = {
            "A": 14, "B": 10, "C": 20, "D": 30, "E": 14, "F": 45,
            "G": 18, "H": 12, "I": 16, "J": 16, "K": 18, "L": 16,
            "M": 20, "N": 14, "O": 14, "P": 40, "Q": 40, "R": 14, "S": 30,
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width
        wb.save(str(p))

    return wb, p


def _find_row_by_app_id(ws, app_id: str) -> Optional[int]:
    """Find the row number for a given app_id (1-indexed, header is row 1)."""
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == app_id:
            return row_idx
    return None


def _add_business_days(start_date: datetime, days: int) -> datetime:
    """Add N business days to a date."""
    current = start_date
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5:  # Monday=0 .. Friday=4
            added += 1
    return current


def upsert_application(app_data: dict, path: Optional[Path] = None) -> str:
    """Insert or update an application row. Returns 'inserted' or 'updated'."""
    wb, p = _ensure_workbook(path)
    ws = wb.active

    app_id = app_data.get("app_id", "")
    existing_row = _find_row_by_app_id(ws, app_id) if app_id else None

    # Compute next_action_due based on status
    status = app_data.get("status", "DISCOVERED")
    next_action = ""
    if status == "SUBMITTED":
        submitted_date = app_data.get("date_submitted")
        if submitted_date:
            try:
                dt = datetime.fromisoformat(submitted_date)
            except (ValueError, TypeError):
                dt = datetime.now(timezone.utc)
            follow_up = _add_business_days(dt, 7)
            next_action = follow_up.strftime("%Y-%m-%d")
    elif status == "NEEDS_HUMAN":
        next_action = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    elif status == "INTERVIEW":
        next_action = _add_business_days(
            datetime.now(timezone.utc), 2
        ).strftime("%Y-%m-%d")

    # Compute completeness from missing_fields if present
    missing_fields_raw = app_data.get("missing_fields", "")
    if missing_fields_raw:
        try:
            missing_list = json.loads(missing_fields_raw) if isinstance(missing_fields_raw, str) else missing_fields_raw
        except (ValueError, TypeError):
            missing_list = []
    else:
        missing_list = []
    missing_count = len(missing_list)

    # proof_path from proof_json
    proof_path = ""
    proof_json_raw = app_data.get("proof_json", "")
    if proof_json_raw:
        try:
            pj = json.loads(proof_json_raw) if isinstance(proof_json_raw, str) else proof_json_raw
            proof_path = pj.get("proof_path", "") if isinstance(pj, dict) else ""
        except (ValueError, TypeError):
            pass

    row_data = {
        "app_id": app_id,
        "source": app_data.get("source", "YC"),
        "company": app_data.get("company", ""),
        "role_title": app_data.get("role_title", ""),
        "role_family": app_data.get("role_family", ""),
        "job_url": app_data.get("job_url", ""),
        "location": app_data.get("location", ""),
        "match_score": app_data.get("match_score", 0),
        "resume_version": app_data.get("resume_version", ""),
        "status": status,
        "policy": app_data.get("policy", "pause_at_submit"),
        "completeness_pct": "" if missing_count == 0 else f"{max(0, 100 - missing_count * 10)}%",
        "missing_fields_count": missing_count if missing_count else "",
        "date_discovered": app_data.get("date_discovered",
                                        datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        "date_submitted": app_data.get("date_submitted", ""),
        "submission_proof": app_data.get("submission_proof", ""),
        "proof_path": proof_path,
        "next_action_due": next_action,
        "notes": app_data.get("notes", ""),
    }

    if existing_row:
        # Update existing row — preserve fields not in app_data
        for col_idx, col_name in enumerate(COLUMNS, 1):
            if col_name in row_data and row_data[col_name]:
                ws.cell(row=existing_row, column=col_idx, value=row_data[col_name])
        # Apply status color
        _apply_status_color(ws, existing_row, status)
        wb.save(str(p))
        return "updated"
    else:
        # Insert new row
        next_row = ws.max_row + 1
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.cell(row=next_row, column=col_idx, value=row_data.get(col_name, ""))
        _apply_status_color(ws, next_row, status)
        wb.save(str(p))
        return "inserted"


def _apply_status_color(ws, row: int, status: str):
    """Apply background color to the status cell based on status."""
    status_col = COLUMNS.index("status") + 1
    color = STATUS_COLORS.get(status, "FFFFFF")
    ws.cell(row=row, column=status_col).fill = PatternFill(
        start_color=color, end_color=color, fill_type="solid"
    )


def validate_transition(current_status: str, new_status: str) -> bool:
    """Check if a status transition is valid."""
    allowed = VALID_TRANSITIONS.get(current_status, [])
    return new_status in allowed


def rebuild_excel(db_path=None, output_path: Optional[Path] = None) -> int:
    """Drop and rebuild applications.xlsx from the DB. Returns row count written."""
    from src.storage.db import get_all_applications_with_jobs, get_connection

    out = output_path or EXCEL_PATH
    out.parent.mkdir(parents=True, exist_ok=True)
    # Delete existing file so _ensure_workbook creates fresh
    if out.exists():
        out.unlink()

    conn = get_connection(db_path)
    apps = get_all_applications_with_jobs(conn)
    conn.close()

    count = 0
    for app in apps:
        # Normalize date fields
        created_at = app.get("created_at", "")
        app_data = {
            **app,
            "date_discovered": created_at[:10] if created_at else "",
            "date_submitted": (app.get("applied_at") or "")[:10],
        }
        upsert_application(app_data, path=out)
        count += 1

    return count


def get_daily_summary(path: Optional[Path] = None) -> dict:
    """Generate a summary of current application statuses."""
    wb, _ = _ensure_workbook(path)
    ws = wb.active

    summary = {
        "total": 0,
        "by_status": {},
        "needs_human": [],
        "upcoming_follow_ups": [],
    }

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    for row_idx in range(2, ws.max_row + 1):
        app_id = ws.cell(row=row_idx, column=1).value
        if not app_id:
            continue

        summary["total"] += 1
        status = ws.cell(row=row_idx, column=COLUMNS.index("status") + 1).value or ""
        company = ws.cell(row=row_idx, column=COLUMNS.index("company") + 1).value or ""
        role = ws.cell(row=row_idx, column=COLUMNS.index("role_title") + 1).value or ""
        next_action = ws.cell(
            row=row_idx, column=COLUMNS.index("next_action_due") + 1
        ).value or ""

        summary["by_status"][status] = summary["by_status"].get(status, 0) + 1

        if status == "NEEDS_HUMAN":
            summary["needs_human"].append({
                "app_id": app_id, "company": company, "role": role
            })

        if next_action and next_action <= today:
            summary["upcoming_follow_ups"].append({
                "app_id": app_id, "company": company, "role": role,
                "due": next_action
            })

    return summary
