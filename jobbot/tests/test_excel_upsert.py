"""Tests for Excel upsert and status tracking."""

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pytest
from openpyxl import load_workbook

from src.storage.excel import (
    COLUMNS,
    _add_business_days,
    get_daily_summary,
    upsert_application,
    validate_transition,
)


@pytest.fixture
def tmp_excel():
    """Create a temporary path for Excel file (must not pre-exist)."""
    path = Path(tempfile.mkdtemp()) / "test_apps.xlsx"
    yield path
    path.unlink(missing_ok=True)


class TestUpsertApplication:
    def test_insert_new_row(self, tmp_excel):
        result = upsert_application({
            "app_id": "test001",
            "company": "Acme Corp",
            "role_title": "Software Engineer",
            "role_family": "fullstack",
            "job_url": "https://boards.greenhouse.io/acme/jobs/1",
            "location": "San Francisco, CA",
            "match_score": 0.85,
            "status": "DISCOVERED",
        }, path=tmp_excel)

        assert result == "inserted"

        wb = load_workbook(str(tmp_excel))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "test001"
        assert ws.cell(row=2, column=3).value == "Acme Corp"
        assert ws.cell(row=2, column=4).value == "Software Engineer"

    def test_update_existing_row(self, tmp_excel):
        # Insert first
        upsert_application({
            "app_id": "test002",
            "company": "Beta Inc",
            "role_title": "ML Engineer",
            "status": "DISCOVERED",
        }, path=tmp_excel)

        # Update
        result = upsert_application({
            "app_id": "test002",
            "status": "SUBMITTED",
            "date_submitted": "2024-01-15",
        }, path=tmp_excel)

        assert result == "updated"

        wb = load_workbook(str(tmp_excel))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "test002"
        status_col = COLUMNS.index("status") + 1
        assert ws.cell(row=2, column=status_col).value == "SUBMITTED"

    def test_multiple_rows(self, tmp_excel):
        for i in range(5):
            upsert_application({
                "app_id": f"multi{i:03d}",
                "company": f"Company {i}",
                "role_title": f"Role {i}",
                "status": "DISCOVERED",
            }, path=tmp_excel)

        wb = load_workbook(str(tmp_excel))
        ws = wb.active
        # 1 header + 5 data rows
        assert ws.max_row == 6

    def test_headers_correct(self, tmp_excel):
        upsert_application({
            "app_id": "hdr001",
            "status": "DISCOVERED",
        }, path=tmp_excel)

        wb = load_workbook(str(tmp_excel))
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, len(COLUMNS) + 1)]
        assert headers == COLUMNS

    def test_submitted_sets_follow_up(self, tmp_excel):
        upsert_application({
            "app_id": "fu001",
            "status": "SUBMITTED",
            "date_submitted": "2024-01-15",
        }, path=tmp_excel)

        wb = load_workbook(str(tmp_excel))
        ws = wb.active
        next_action_col = COLUMNS.index("next_action_due") + 1
        next_action = ws.cell(row=2, column=next_action_col).value
        assert next_action is not None
        assert next_action != ""

    def test_needs_human_sets_today(self, tmp_excel):
        upsert_application({
            "app_id": "nh001",
            "status": "NEEDS_HUMAN",
        }, path=tmp_excel)

        wb = load_workbook(str(tmp_excel))
        ws = wb.active
        next_action_col = COLUMNS.index("next_action_due") + 1
        next_action = ws.cell(row=2, column=next_action_col).value
        assert next_action is not None


class TestValidateTransition:
    def test_discovered_to_ready(self):
        assert validate_transition("DISCOVERED", "READY_TO_APPLY") is True

    def test_ready_to_applying(self):
        assert validate_transition("READY_TO_APPLY", "APPLYING") is True

    def test_applying_to_submitted(self):
        assert validate_transition("APPLYING", "SUBMITTED") is True

    def test_applying_to_needs_human(self):
        assert validate_transition("APPLYING", "NEEDS_HUMAN") is True

    def test_submitted_to_interview(self):
        assert validate_transition("SUBMITTED", "INTERVIEW") is True

    def test_needs_human_to_applying(self):
        assert validate_transition("NEEDS_HUMAN", "APPLYING") is True

    def test_invalid_discovered_to_submitted(self):
        assert validate_transition("DISCOVERED", "SUBMITTED") is False

    def test_rejected_no_transitions(self):
        assert validate_transition("REJECTED", "DISCOVERED") is False


class TestAddBusinessDays:
    def test_add_5_business_days(self):
        from datetime import datetime
        start = datetime(2024, 1, 15)  # Monday
        result = _add_business_days(start, 5)
        assert result.weekday() == 0  # Next Monday
        assert result.day == 22

    def test_add_7_business_days_over_weekend(self):
        from datetime import datetime
        start = datetime(2024, 1, 15)  # Monday
        result = _add_business_days(start, 7)
        assert result.weekday() == 2  # Wednesday
        assert result.day == 24

    def test_skip_weekends(self):
        from datetime import datetime
        start = datetime(2024, 1, 19)  # Friday
        result = _add_business_days(start, 1)
        assert result.weekday() == 0  # Monday
        assert result.day == 22


class TestDailySummary:
    def test_empty_summary(self, tmp_excel):
        summary = get_daily_summary(tmp_excel)
        assert summary["total"] == 0
        assert summary["by_status"] == {}

    def test_summary_with_data(self, tmp_excel):
        upsert_application({
            "app_id": "s001", "status": "SUBMITTED", "company": "A", "role_title": "SWE",
        }, path=tmp_excel)
        upsert_application({
            "app_id": "s002", "status": "NEEDS_HUMAN", "company": "B", "role_title": "ML",
        }, path=tmp_excel)
        upsert_application({
            "app_id": "s003", "status": "SUBMITTED", "company": "C", "role_title": "FE",
        }, path=tmp_excel)

        summary = get_daily_summary(tmp_excel)
        assert summary["total"] == 3
        assert summary["by_status"]["SUBMITTED"] == 2
        assert summary["by_status"]["NEEDS_HUMAN"] == 1
        assert len(summary["needs_human"]) == 1
